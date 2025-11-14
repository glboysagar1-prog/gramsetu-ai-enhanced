"""
GramSetu AI - National Governance Intelligence Network
Backend API for Citizen Complaint Management System

Features:
- Level 1: Citizen IVR simulation with text input
- Level 3: Complaint Validation & CRS (Citizen Rating System)
- Level 4: Blockchain Audit simulation with SHA256 hashing
- NLP: Zero-shot classification and duplicate detection
- Multilingual support (English for now, ready for IndicASR)
"""

import os
import sqlite3
import hashlib
import json
import logging
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from werkzeug.utils import secure_filename
from functools import wraps
from io import BytesIO
import time
import signal

from flask import Flask, request, jsonify, Response, send_file
from flask_cors import CORS

# Optional AI imports - graceful degradation
try:
    from transformers import pipeline
    from sentence_transformers import SentenceTransformer
    AI_AVAILABLE = True
except ImportError:
    pipeline = None
    SentenceTransformer = None
    AI_AVAILABLE = False
    print("⚠️  AI libraries not available - using fallback mode")

try:
    import numpy as np
    from sklearn.metrics.pairwise import cosine_similarity
except ImportError:
    np = None
    cosine_similarity = None
    print("⚠️  NumPy/sklearn not available - using basic mode")

# Import voice complaint services (optional)
try:
    from services.voice_complaint_service import get_voice_service
    from services.multilingual_classifier import get_classifier
    VOICE_SERVICES_AVAILABLE = True
except ImportError:
    get_voice_service = None
    get_classifier = None
    VOICE_SERVICES_AVAILABLE = False
    print("⚠️  Voice services not available - using text-only mode")

# Initialize Flask app
app = Flask(__name__)
CORS(app)  # Enable CORS for all routes

# API version
API_VERSION = 'v1'

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# Add file handler
file_handler = logging.FileHandler("gramsetu.log")
logging.getLogger('').addHandler(file_handler)
logger = logging.getLogger(__name__)

# Global variables for AI models (loaded once at startup)
zero_shot_classifier = None
sentence_model = None
voice_service = None
multilingual_classifier = None

# Upload configuration
UPLOAD_FOLDER = 'uploads/audio'
ALLOWED_EXTENSIONS = {'wav', 'mp3', 'ogg', 'm4a', 'flac', 'webm'}
MAX_AUDIO_SIZE = 10 * 1024 * 1024  # 10 MB

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Database configuration
DB_PATH = 'gramsetu_ai.db'

# API version
API_VERSION = 'v1'

# Complaint categories for zero-shot classification
COMPLAINT_CATEGORIES = [
    "Water supply issues",
    "Health and medical services", 
    "Electricity and power problems",
    "Road and infrastructure",
    "Other government services"
]

# Invalid context patterns (spam/irrelevant detection)
INVALID_PATTERNS = [
    "rain not coming", "weather", "cricket", "movie", "food delivery",
    "shopping", "entertainment", "personal relationship"
]

def initialize_ai_models():
    """Initialize Hugging Face models for NLP processing (Lazy Loading)"""
    global zero_shot_classifier, sentence_model, voice_service, multilingual_classifier
    
    print("AI Models will be loaded on-demand (lazy loading for faster startup)")
    
    # Set to None - models will load when first requested
    zero_shot_classifier = None
    sentence_model = None
    
    if not AI_AVAILABLE:
        print("⚠️  AI libraries not installed - running in fallback mode")
        print("✓ System will use mock responses for demo")
        return
    
    print("✅ AI libraries available - models will load on first use")
    
    # Voice complaint service (lazy loading for better startup time)
    # Multilingual classifier
    if VOICE_SERVICES_AVAILABLE and get_classifier:
        try:
            multilingual_classifier = get_classifier()
            print("Multilingual classifier loaded!")
        except Exception as e:
            print(f"Warning: Multilingual classifier not loaded: {e}")
    else:
        print("⚠️  Voice services not available - skipping classifier")
        multilingual_classifier = None

def load_zero_shot_classifier():
    """Lazy load zero-shot classifier on first use"""
    global zero_shot_classifier
    if zero_shot_classifier is None and AI_AVAILABLE:
        try:
            print("Loading zero-shot classifier (distilbart-mnli-12-3)...")
            zero_shot_classifier = pipeline(
                "zero-shot-classification",
                model="valhalla/distilbart-mnli-12-3",
                device=-1
            )
            print("✓ Zero-shot classifier loaded!")
        except Exception as e:
            print(f"Error loading classifier: {e}")
    return zero_shot_classifier

def load_sentence_model():
    """Lazy load sentence transformer on first use"""
    global sentence_model
    if sentence_model is None and AI_AVAILABLE:
        try:
            print("Loading sentence transformer (MiniLM-L6-v2)...")
            sentence_model = SentenceTransformer('all-MiniLM-L6-v2')
            print("✓ Sentence transformer loaded!")
        except Exception as e:
            print(f"Error loading sentence model: {e}")
    return sentence_model

def init_database():
    """Initialize SQLite database with required tables"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Citizens table for CRS tracking
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS citizens (
            id TEXT PRIMARY KEY,
            crs_score INTEGER DEFAULT 100,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Complaints table with blockchain hash
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS complaints (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            text TEXT NOT NULL,
            category TEXT,
            urgency TEXT DEFAULT 'Medium',
            citizen_id TEXT NOT NULL,
            crs_score INTEGER DEFAULT 100,
            hash TEXT UNIQUE NOT NULL,
            timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            status TEXT DEFAULT 'Pending',
            evidence TEXT,
            is_duplicate BOOLEAN DEFAULT FALSE,
            is_valid BOOLEAN DEFAULT TRUE,
            FOREIGN KEY (citizen_id) REFERENCES citizens (id)
        )
    ''')
    
    # Create indexes for better performance
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_citizen_id ON complaints(citizen_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_hash ON complaints(hash)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_status ON complaints(status)')
    
    # Create field_workers table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS field_workers (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            area TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Create assignments table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            complaint_id INTEGER NOT NULL,
            field_worker_id TEXT NOT NULL,
            assigned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            resolved_at TIMESTAMP,
            resolution_notes TEXT,
            FOREIGN KEY (complaint_id) REFERENCES complaints (id),
            FOREIGN KEY (field_worker_id) REFERENCES field_workers (id)
        )
    ''')
    
    # Insert sample field workers if not exist
    cursor.execute('SELECT COUNT(*) FROM field_workers')
    if cursor.fetchone()[0] == 0:
        sample_workers = [
            ('FW001', 'Rajesh Kumar', 'North Zone'),
            ('FW002', 'Priya Singh', 'South Zone'),
            ('FW003', 'Amit Sharma', 'East Zone'),
            ('FW004', 'Deepa Patel', 'West Zone')
        ]
        cursor.executemany('INSERT INTO field_workers (id, name, area) VALUES (?, ?, ?)', sample_workers)
    
    conn.commit()
    conn.close()
    logger.info("Database initialized successfully!")

def generate_blockchain_hash(text: str, timestamp: str) -> str:
    """Generate SHA256 hash for blockchain simulation"""
    data = f"{text}{timestamp}"
    return hashlib.sha256(data.encode()).hexdigest()

def validate_complaint_context(text: str) -> Tuple[bool, str]:
    """Validate if complaint is relevant to governance services"""
    text_lower = text.lower()
    
    # Check for invalid patterns
    for pattern in INVALID_PATTERNS:
        if pattern in text_lower:
            return False, f"Invalid context: '{pattern}' detected"
    
    # Check minimum length
    if len(text.strip()) < 10:
        return False, "Complaint too short (minimum 10 characters)"
    
    return True, "Valid context"

def classify_complaint(text: str) -> str:
    """Classify complaint using zero-shot classification"""
    classifier = load_zero_shot_classifier()  # Lazy load
    
    if not classifier:
        # Fallback: simple keyword matching
        text_lower = text.lower()
        if any(word in text_lower for word in ['water', 'tap', 'supply', 'नल', 'पानी']):
            return "Water supply issues"
        elif any(word in text_lower for word in ['electricity', 'power', 'light', 'बिजली']):
            return "Electricity and power problems"
        elif any(word in text_lower for word in ['road', 'pothole', 'सड़क']):
            return "Road and infrastructure"
        elif any(word in text_lower for word in ['hospital', 'doctor', 'health', 'अस्पताल']):
            return "Health and medical services"
        return "Other government services"
    
    try:
        result = classifier(text, COMPLAINT_CATEGORIES)
        return result['labels'][0]  # Return top category
    except Exception as e:
        print(f"Classification error: {e}")
        return "Other government services"

def detect_urgency(text: str) -> str:
    """Detect urgency level based on keywords"""
    urgent_keywords = ['urgent', 'emergency', 'critical', 'immediate', 'asap']
    text_lower = text.lower()
    
    if any(keyword in text_lower for keyword in urgent_keywords):
        return "High"
    return "Medium"

def detect_duplicates(text: str, citizen_id: str) -> Tuple[bool, Optional[int]]:
    """Detect duplicate complaints using sentence transformers"""
    s_model = load_sentence_model()  # Lazy load
    
    if not s_model or not cosine_similarity:
        # Fallback: simple text matching
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT id, text FROM complaints 
                WHERE citizen_id = ? AND timestamp > datetime('now', '-30 days')
            ''', (citizen_id,))
            recent_complaints = cursor.fetchall()
            conn.close()
            
            for complaint_id, complaint_text in recent_complaints:
                if text.lower() == complaint_text.lower():
                    return True, complaint_id
            return False, None
        except Exception as e:
            print(f"Duplicate detection error: {e}")
            return False, None
    
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Get recent complaints from same citizen (last 30 days)
        cursor.execute('''
            SELECT id, text FROM complaints 
            WHERE citizen_id = ? AND timestamp > datetime('now', '-30 days')
        ''', (citizen_id,))
        
        recent_complaints = cursor.fetchall()
        conn.close()
        
        if not recent_complaints:
            return False, None
        
        # Encode current complaint
        current_embedding = s_model.encode([text])
        
        # Check similarity with recent complaints
        for complaint_id, complaint_text in recent_complaints:
            complaint_embedding = s_model.encode([complaint_text])
            similarity = cosine_similarity(current_embedding, complaint_embedding)[0][0]
            
            if similarity > 0.9:  # High similarity threshold
                return True, complaint_id
        
        return False, None
        
    except Exception as e:
        print(f"Duplicate detection error: {e}")
        return False, None

def update_citizen_crs(citizen_id: str, is_valid: bool, is_duplicate: bool):
    """Update Citizen Rating System score"""
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Get current CRS score
    cursor.execute('SELECT crs_score FROM citizens WHERE id = ?', (citizen_id,))
    result = cursor.fetchone()
    
    if result:
        current_score = result[0]
    else:
        # Create new citizen with default score
        current_score = 100
        cursor.execute('INSERT INTO citizens (id, crs_score) VALUES (?, ?)', (citizen_id, current_score))
    
    # Update score based on validation
    if not is_valid:
        current_score = max(0, current_score - 10)
    elif is_duplicate:
        current_score = max(0, current_score - 5)
    else:
        current_score = min(100, current_score + 1)
    
    cursor.execute('UPDATE citizens SET crs_score = ? WHERE id = ?', (current_score, citizen_id))
    conn.commit()
    conn.close()

# ============================================
# BACKEND INTEGRATIONS - Production Ready
# ============================================

# Load AI fallbacks
try:
    with open('mocks/ai_fallbacks.json', 'r') as f:
        AI_FALLBACKS = json.load(f)
    logger.info("AI fallbacks loaded successfully")
except:
    AI_FALLBACKS = {}
    logger.warning("AI fallbacks not loaded")

# Redis Cache Integration
redis_client = None
try:
    import redis
    if os.getenv('REDIS_URL'):
        redis_client = redis.from_url(os.getenv('REDIS_URL'))
        redis_client.ping()
        logger.info("✓ Redis connected successfully")
except:
    logger.warning("Redis not available, using in-memory cache")

# In-memory cache fallback
in_memory_cache = {}

def cache_response(ttl=10):
    """
    Decorator to cache API responses
    Args:
        ttl: Time to live in seconds
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # Generate cache key
            cache_key = f"{request.path}:{json.dumps(request.args.to_dict())}"
            
            # Try to get from cache
            if redis_client:
                try:
                    cached = redis_client.get(cache_key)
                    if cached:
                        logger.debug(f"✓ Cache hit: {cache_key}")
                        return jsonify(json.loads(cached))
                except:
                    pass
            elif cache_key in in_memory_cache:
                cached_data, expires_at = in_memory_cache[cache_key]
                if time.time() < expires_at:
                    logger.debug(f"✓ Memory cache hit: {cache_key}")
                    return jsonify(cached_data)
            
            # Cache miss - execute function
            response = f(*args, **kwargs)
            response_data = response.get_json()
            
            # Store in cache
            if redis_client:
                try:
                    redis_client.setex(cache_key, ttl, json.dumps(response_data))
                except:
                    pass
            else:
                in_memory_cache[cache_key] = (response_data, time.time() + ttl)
            
            return response
        return decorated_function
    return decorator

def api_call_wrapper(timeout=8, max_retries=1, fallback=None):
    """
    Wrapper for external API calls with timeout, retry, and fallback
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            for attempt in range(max_retries + 1):
                try:
                    # Set timeout using signal
                    def timeout_handler(signum, frame):
                        raise TimeoutError("API call exceeded timeout")
                    
                    signal.signal(signal.SIGALRM, timeout_handler)
                    signal.alarm(timeout)
                    
                    result = f(*args, **kwargs)
                    signal.alarm(0)
                    return result
                    
                except (TimeoutError, Exception) as e:
                    signal.alarm(0)
                    logger.error(f"API call failed (attempt {attempt + 1}/{max_retries + 1}): {str(e)}")
                    
                    if attempt < max_retries:
                        time.sleep(1)  # Wait before retry
                        continue
                    
                    # Log to audit
                    log_audit_event('api_failure', {
                        'function': f.__name__,
                        'error': str(e),
                        'fallback_used': fallback is not None
                    })
                    
                    # Use fallback if available
                    if fallback:
                        return fallback(*args, **kwargs)
                    raise
        
        return decorated_function
    return decorator

def log_audit_event(event_type: str, data: dict):
    """Log audit events to file"""
    try:
        os.makedirs('logs', exist_ok=True)
        with open('logs/audit.log', 'a') as f:
            f.write(json.dumps({
                'timestamp': datetime.utcnow().isoformat(),
                'event': event_type,
                'data': data
            }) + '\n')
    except:
        pass

# OpenAI Integration with Fallback
def get_ai_response(query: str, role: str) -> str:
    """
    Get AI response with automatic fallback
    """
    # Check if OpenAI is configured
    if not os.getenv('OPENAI_API_KEY'):
        logger.info("Using AI fallback (no API key)")
        return get_fallback_ai_response(query, role)
    
    try:
        import openai
        openai.api_key = os.getenv('OPENAI_API_KEY')
        
        # Role-specific system prompts
        system_prompts = {
            'citizen': "You are a helpful governance assistant helping citizens track complaints and understand government processes. Provide clear, concise answers.",
            'field': "You are a field operations assistant helping workers optimize routes and prioritize tasks. Focus on actionable advice.",
            'district': "You are a district analytics assistant providing insights on ward performance and resource allocation. Use data-driven recommendations.",
            'state': "You are a state policy assistant analyzing integrity trends and district comparisons. Provide strategic insights.",
            'national': "You are a national governance strategist providing comparative state analytics and policy recommendations. Think at scale."
        }
        
        @api_call_wrapper(timeout=8, max_retries=1, fallback=lambda q, r: get_fallback_ai_response(q, r))
        def call_openai(prompt, role_key):
            response = openai.ChatCompletion.create(
                model=os.getenv('OPENAI_MODEL', 'gpt-3.5-turbo'),
                messages=[
                    {"role": "system", "content": system_prompts.get(role_key, system_prompts['citizen'])},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=int(os.getenv('OPENAI_MAX_TOKENS', 500)),
                temperature=float(os.getenv('OPENAI_TEMPERATURE', 0.7))
            )
            return response.choices[0].message.content
        
        return call_openai(query, role)
        
    except Exception as e:
        logger.error(f"OpenAI API failed: {str(e)}")
        return get_fallback_ai_response(query, role)

def get_fallback_ai_response(query: str, role: str) -> str:
    """
    Get response from fallback JSON
    """
    query_lower = query.lower()
    
    if not AI_FALLBACKS:
        return "I'm here to help! Could you please rephrase your question?"
    
    role_responses = AI_FALLBACKS.get('openai', {}).get('governance_queries', {})
    
    # Match keywords
    for key, responses in role_responses.items():
        if isinstance(responses, dict) and key in query_lower:
            return responses.get(role, responses.get('default', ''))
    
    # Default response
    return role_responses.get('default', "I can help with complaint tracking, analytics, and governance insights. What would you like to know?")

# Blockchain Integration with Mock Fallback
def log_to_blockchain(complaint_id: str, complaint_data: dict) -> dict:
    """
    Log complaint to blockchain with automatic mock fallback
    """
    if not os.getenv('THIRDWEB_SECRET_KEY'):
        logger.info("Using blockchain mock (no API key)")
        return generate_mock_blockchain_tx(complaint_id, complaint_data)
    
    try:
        from web3 import Web3
        
        # Initialize Web3
        provider_url = os.getenv('WEB3_PROVIDER_URL', 'https://polygon-rpc.com')
        w3 = Web3(Web3.HTTPProvider(provider_url))
        
        # Generate transaction hash
        tx_data = f"{complaint_id}{json.dumps(complaint_data)}"
        tx_hash = '0x' + hashlib.sha256(tx_data.encode()).hexdigest()
        
        # In production, this would interact with smart contract
        # For now, simulate successful transaction
        
        return {
            'tx_hash': tx_hash,
            'block_number': w3.eth.block_number if w3.isConnected() else 45000000,
            'explorer_url': f"https://polygonscan.com/tx/{tx_hash}",
            'status': 'confirmed',
            'gas_used': '21000',
            'timestamp': datetime.utcnow().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Blockchain logging failed: {str(e)}")
        return generate_mock_blockchain_tx(complaint_id, complaint_data)

def generate_mock_blockchain_tx(complaint_id: str, complaint_data: dict) -> dict:
    """
    Generate deterministic mock transaction
    """
    tx_data = f"{complaint_id}{json.dumps(complaint_data)}"
    tx_hash = 'MCK-TX-' + hashlib.sha256(tx_data.encode()).hexdigest()[:40]
    
    # Extract complaint ID number
    try:
        id_num = int(''.join(filter(str.isdigit, str(complaint_id))))
    except:
        id_num = hash(complaint_id) % 1000000
    
    return {
        'tx_hash': tx_hash,
        'block_number': 45000000 + id_num,
        'explorer_url': f"https://mock.polygonscan.com/tx/{tx_hash}",
        'status': 'mocked',
        'gas_used': '21000',
        'timestamp': datetime.utcnow().isoformat()
    }
    
    
# API Endpoints for Integration
@app.route(f'/api/{API_VERSION}/complaints', methods=['GET'])
def get_complaints():
    """Get all complaints for dashboard"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get all complaints with assignment info
        cursor.execute('''
            SELECT 
                c.id, c.text, c.category, c.urgency, c.citizen_id, 
                c.status, c.timestamp, c.hash, c.is_valid, c.is_duplicate,
                a.field_worker_id, fw.name as field_worker_name, 
                a.assigned_at, a.resolved_at, a.resolution_notes
            FROM complaints c
            LEFT JOIN assignments a ON c.id = a.complaint_id
            LEFT JOIN field_workers fw ON a.field_worker_id = fw.id
            ORDER BY c.timestamp DESC
        ''')
        
        complaints = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        logger.info(f"Retrieved {len(complaints)} complaints")
        return jsonify({"status": "success", "data": complaints})
    
    except Exception as e:
        logger.error(f"Error retrieving complaints: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route(f'/api/{API_VERSION}/complaints', methods=['POST'])
def submit_complaint():
    """Submit a new complaint from Flutter app"""
    try:
        data = request.json
        
        if not data or 'text' not in data or 'citizen_id' not in data:
            return jsonify({"status": "error", "message": "Missing required fields"}), 400
        
        text = data['text']
        citizen_id = data['citizen_id']
        
        # Validate complaint context
        is_valid, validation_message = validate_complaint_context(text)
        
        # Check for duplicates
        is_duplicate, duplicate_id = detect_duplicates(text, citizen_id)
        
        # Generate timestamp and hash
        timestamp = datetime.now().isoformat()
        hash_value = generate_blockchain_hash(text, timestamp)
        
        # Classify complaint
        category = classify_complaint(text)
        
        # Detect urgency
        urgency = detect_urgency(text)
        
        # Update CRS score
        update_citizen_crs(citizen_id, is_valid, is_duplicate)
        
        # Save complaint to database
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO complaints 
            (text, category, urgency, citizen_id, hash, is_valid, is_duplicate, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (text, category, urgency, citizen_id, hash_value, is_valid, is_duplicate, 
              "Invalid" if not is_valid else "Duplicate" if is_duplicate else "Pending"))
        
        complaint_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        logger.info(f"New complaint submitted: ID={complaint_id}, Citizen={citizen_id}")
        
        return jsonify({
            "status": "success", 
            "data": {
                "complaint_id": complaint_id,
                "hash": hash_value,
                "is_valid": is_valid,
                "validation_message": validation_message,
                "is_duplicate": is_duplicate,
                "duplicate_id": duplicate_id,
                "category": category,
                "urgency": urgency
            }
        })
        
    except Exception as e:
        logger.error(f"Error submitting complaint: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route(f'/api/{API_VERSION}/complaints/<int:complaint_id>', methods=['PUT'])
def update_complaint(complaint_id):
    """Update complaint status (from field worker app)"""
    try:
        data = request.json
        
        if not data:
            return jsonify({"status": "error", "message": "No data provided"}), 400
        
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Check if complaint exists
        cursor.execute('SELECT id FROM complaints WHERE id = ?', (complaint_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({"status": "error", "message": "Complaint not found"}), 404
        
        # Update status if provided
        if 'status' in data:
            cursor.execute('UPDATE complaints SET status = ? WHERE id = ?', 
                          (data['status'], complaint_id))
        
        # Handle assignment
        if 'field_worker_id' in data:
            # Check if already assigned
            cursor.execute('SELECT id FROM assignments WHERE complaint_id = ?', (complaint_id,))
            assignment = cursor.fetchone()
            
            if assignment:
                # Update existing assignment
                cursor.execute('''
                    UPDATE assignments 
                    SET field_worker_id = ?, 
                        resolved_at = CASE WHEN ? = 'Resolved' THEN CURRENT_TIMESTAMP ELSE NULL END,
                        resolution_notes = ?
                    WHERE complaint_id = ?
                ''', (data['field_worker_id'], 
                      data.get('status', ''), 
                      data.get('resolution_notes', ''),
                      complaint_id))
            else:
                # Create new assignment
                cursor.execute('''
                    INSERT INTO assignments 
                    (complaint_id, field_worker_id, resolved_at, resolution_notes)
                    VALUES (?, ?, ?, ?)
                ''', (complaint_id, 
                      data['field_worker_id'],
                      datetime.now().isoformat() if data.get('status') == 'Resolved' else None,
                      data.get('resolution_notes', '')))
        
        conn.commit()
        conn.close()
        
        logger.info(f"Complaint {complaint_id} updated: {data}")
        return jsonify({"status": "success"})
        
    except Exception as e:
        logger.error(f"Error updating complaint: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route(f'/api/{API_VERSION}/field-workers', methods=['GET'])
def get_field_workers():
    """Get all field workers"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM field_workers')
        workers = [dict(row) for row in cursor.fetchall()]
        conn.close()
        
        return jsonify({"status": "success", "data": workers})
    
    except Exception as e:
        logger.error(f"Error retrieving field workers: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

# ============================================
# NEW PRODUCTION ENDPOINTS
# ============================================

@app.route('/healthz', methods=['GET'])
def health_check():
    """
    Health check endpoint for monitoring and load balancers
    """
    health_status = {
        'status': 'healthy',
        'timestamp': datetime.utcnow().isoformat(),
        'version': '1.0.0',
        'checks': {}
    }
    
    # Check database
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute('SELECT 1')
        cursor.close()
        conn.close()
        health_status['checks']['database'] = 'ok'
    except Exception as e:
        health_status['checks']['database'] = f'error: {str(e)}'
        health_status['status'] = 'unhealthy'
    
    # Check Redis (if configured)
    if redis_client:
        try:
            redis_client.ping()
            health_status['checks']['redis'] = 'ok'
        except Exception as e:
            health_status['checks']['redis'] = f'error: {str(e)}'
    else:
        health_status['checks']['redis'] = 'not_configured'
    
    # Check external APIs
    health_status['checks']['openai'] = 'mock' if not os.getenv('OPENAI_API_KEY') else 'configured'
    health_status['checks']['thirdweb'] = 'mock' if not os.getenv('THIRDWEB_SECRET_KEY') else 'configured'
    
    status_code = 200 if health_status['status'] == 'healthy' else 503
    return jsonify(health_status), status_code

@app.route('/api/dashboard', methods=['GET'])
@cache_response(ttl=10)
def get_dashboard():
    """
    Get dashboard data for specific role (with caching)
    """
    role = request.args.get('role', 'citizen')
    time_range = request.args.get('timeRange', '30d')
    
    try:
        # Load seed data for offline/demo mode
        if os.getenv('OFFLINE_MODE', 'false').lower() == 'true':
            with open('demo/seed_data.json', 'r') as f:
                seed_data = json.load(f)
                return jsonify({"status": "success", "data": seed_data})
        
        # Real data from database
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        dashboard_data = {
            'total_complaints': 0,
            'resolved': 0,
            'pending': 0,
            'in_progress': 0
        }
        
        # Get complaint statistics
        cursor.execute("SELECT COUNT(*) as count, status FROM complaints GROUP BY status")
        for row in cursor.fetchall():
            status = row['status']
            count = row['count']
            dashboard_data['total_complaints'] += count
            if status == 'Resolved':
                dashboard_data['resolved'] = count
            elif status == 'Pending':
                dashboard_data['pending'] = count
            elif status == 'In Progress':
                dashboard_data['in_progress'] = count
        
        conn.close()
        
        return jsonify({"status": "success", "data": dashboard_data})
        
    except Exception as e:
        logger.error(f"Error getting dashboard data: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/ai/chat', methods=['POST'])
def ai_chat():
    """
    AI chat endpoint with OpenAI integration and fallback
    """
    try:
        data = request.get_json()
        query = data.get('message', '')
        role = data.get('role', 'citizen')
        
        if not query:
            return jsonify({"status": "error", "message": "No message provided"}), 400
        
        # Get AI response (with automatic fallback)
        response_text = get_ai_response(query, role)
        
        return jsonify({
            'status': 'success',
            'response': response_text,
            'timestamp': datetime.utcnow().isoformat(),
            'source': 'openai' if os.getenv('OPENAI_API_KEY') else 'fallback'
        })
        
    except Exception as e:
        logger.error(f"Error in AI chat: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/blockchain/log', methods=['POST'])
def blockchain_log():
    """
    Log complaint to blockchain
    """
    try:
        data = request.get_json()
        complaint_id = data.get('complaint_id')
        complaint_data = data.get('data', {})
        
        if not complaint_id:
            return jsonify({"status": "error", "message": "Complaint ID required"}), 400
        
        # Log to blockchain (with automatic mock fallback)
        tx_result = log_to_blockchain(complaint_id, complaint_data)
        
        return jsonify({
            'status': 'success',
            'blockchain': tx_result
        })
        
    except Exception as e:
        logger.error(f"Error logging to blockchain: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route(f'/api/{API_VERSION}/dashboard', methods=['GET'])
def get_dashboard_data():
    """Get dashboard statistics for React dashboard"""
    try:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get complaint statistics
        cursor.execute('''
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
                SUM(CASE WHEN status = 'Resolved' THEN 1 ELSE 0 END) as resolved,
                SUM(CASE WHEN status = 'Invalid' THEN 1 ELSE 0 END) as invalid,
                SUM(CASE WHEN status = 'Duplicate' THEN 1 ELSE 0 END) as duplicate
            FROM complaints
        ''')
        stats = dict(cursor.fetchone())
        
        # Get category distribution
        cursor.execute('''
            SELECT category, COUNT(*) as count
            FROM complaints
            GROUP BY category
            ORDER BY count DESC
        ''')
        categories = [dict(row) for row in cursor.fetchall()]
        
        # Get field worker performance
        cursor.execute('''
            SELECT 
                fw.id, fw.name, fw.area,
                COUNT(a.id) as assigned_count,
                SUM(CASE WHEN a.resolved_at IS NOT NULL THEN 1 ELSE 0 END) as resolved_count
            FROM field_workers fw
            LEFT JOIN assignments a ON fw.id = a.field_worker_id
            GROUP BY fw.id
            ORDER BY resolved_count DESC
        ''')
        workers_performance = [dict(row) for row in cursor.fetchall()]
        
        # Get complaints requiring escalation (older than 72 hours and not resolved)
        cursor.execute('''
            SELECT id, text, category, status, timestamp
            FROM complaints
            WHERE 
                status NOT IN ('Resolved', 'Invalid', 'Duplicate') 
                AND datetime(timestamp) < datetime('now', '-3 days')
            ORDER BY timestamp ASC
        ''')
        escalations = [dict(row) for row in cursor.fetchall()]
        
        conn.close()
        
        return jsonify({
            "status": "success", 
            "data": {
                "stats": stats,
                "categories": categories,
                "workers_performance": workers_performance,
                "escalations": escalations
            }
        })
    
    except Exception as e:
        logger.error(f"Error retrieving dashboard data: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

# Main entry point
if __name__ == '__main__':
    # Initialize database
    init_database()
    
    # Initialize AI models
    initialize_ai_models()
    
    # Start Flask server
    port = int(os.getenv('PORT', 5001))
    logger.info(f"Starting GramSetu AI backend server on http://localhost:{port}")
    app.run(host='0.0.0.0', port=port, debug=True)

# API endpoints for integration
@app.route('/api/v1/complaints', methods=['GET'])
def get_complaints():
    """Get all complaints for dashboard"""
    conn = sqlite3.connect('gramsetu_ai.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM complaints')
    complaints = cursor.fetchall()
    conn.close()
    
    # Convert to JSON
    result = []
    for c in complaints:
        result.append({
            'id': c[0],
            'citizen_id': c[1],
            'description': c[2],
            'category': c[3],
            'status': c[4],
            'created_at': c[5],
            'hash': c[6],
            'urgency': c[7] if len(c) > 7 else 'Medium'
        })
    
    return jsonify(result)

@app.route('/api/v1/complaints', methods=['POST'])
def create_complaint():
    """Create a new complaint with AI processing"""
    data = request.json
    
    # Process complaint
    complaint_text = data.get('text', '')
    citizen_id = data.get('citizenId', 1)
    
    # Validate complaint
    is_valid = validate_complaint_context(complaint_text)
    
    # Classify complaint
    category = classify_complaint(complaint_text)
    
    # Detect urgency
    urgency = detect_urgency(complaint_text)
    
    # Generate hash
    hash_value = generate_blockchain_hash(complaint_text)
    
    # Update CRS score
    update_crs_score(citizen_id, is_valid)
    
    # Save to database
    conn = sqlite3.connect('gramsetu_ai.db')
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO complaints (citizen_id, description, category, status, created_at, hash, urgency)
        VALUES (?, ?, ?, ?, datetime('now'), ?, ?)
    ''', (citizen_id, complaint_text, category, 'Pending' if is_valid else 'Invalid', hash_value, urgency))
    
    complaint_id = cursor.lastrowid
    conn.commit()
    conn.close()
    
    return jsonify({
        'status': 'success',
        'data': {
            'complaint_id': complaint_id,
            'category': category,
            'is_valid': is_valid,
            'hash': hash_value,
            'urgency': urgency
        }
    })

@app.route('/api/v1/complaints/<int:complaint_id>', methods=['PUT'])
def update_complaint(complaint_id):
    """Update a complaint status"""
    data = request.json
    
    conn = sqlite3.connect('gramsetu_ai.db')
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE complaints
        SET status = ?
        WHERE id = ?
    ''', (data.get('status', 'Pending'), complaint_id))
    
    conn.commit()
    conn.close()
    
    return jsonify({
        'status': 'success',
        'message': f'Complaint {complaint_id} updated'
    })

@app.route(f'/api/{API_VERSION}/submit_complaint', methods=['POST'])
def submit_complaint():
    """Submit a new complaint with AI processing"""
    try:
        data = request.get_json()
        
        if not data or 'text' not in data or 'citizen_id' not in data:
            return jsonify({'error': 'Missing required fields: text, citizen_id'}), 400
        
        text = data['text'].strip()
        citizen_id = data['citizen_id'].strip()
        
        # Validate complaint context
        is_valid_context, validation_message = validate_complaint_context(text)
        
        # Detect duplicates
        is_duplicate, duplicate_id = detect_duplicates(text, citizen_id)
        
        # Generate blockchain hash
        timestamp = datetime.now().isoformat()
        complaint_hash = generate_blockchain_hash(text, timestamp)
        
        # AI processing
        category = classify_complaint(text)
        urgency = detect_urgency(text)
        
        # Determine final validation status
        is_valid = is_valid_context and not is_duplicate
        
        # Update CRS score
        crs_score = update_citizen_crs(citizen_id, is_valid, is_duplicate)
        
        # Store complaint in database
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO complaints 
            (text, category, urgency, citizen_id, crs_score, hash, timestamp, 
             status, is_duplicate, is_valid)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (text, category, urgency, citizen_id, crs_score, complaint_hash, 
              timestamp, 'Pending', is_duplicate, is_valid))
        
        complaint_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        # Prepare response
        response = {
            'complaint_id': complaint_id,
            'category': category,
            'urgency': urgency,
            'crs_score': crs_score,
            'hash': complaint_hash,
            'is_valid': is_valid,
            'is_duplicate': is_duplicate,
            'validation_message': validation_message,
            'timestamp': timestamp
        }
        
        if is_duplicate:
            response['duplicate_of'] = duplicate_id
        
        return jsonify(response), 201
        
    except Exception as e:
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

@app.route('/update_complaint', methods=['POST'])
def update_complaint():
    """Update complaint with evidence and status"""
    try:
        data = request.get_json()
        
        if not data or 'id' not in data:
            return jsonify({'error': 'Missing required field: id'}), 400
        
        complaint_id = data['id']
        evidence = data.get('evidence', '')
        status = data.get('status', 'Pending')
        
        # Validate status
        valid_statuses = ['Pending', 'In Progress', 'Resolved', 'Rejected']
        if status not in valid_statuses:
            return jsonify({'error': f'Invalid status. Must be one of: {valid_statuses}'}), 400
        
        # Update complaint in database
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Get current complaint data
        cursor.execute('SELECT citizen_id, is_valid FROM complaints WHERE id = ?', (complaint_id,))
        result = cursor.fetchone()
        
        if not result:
            conn.close()
            return jsonify({'error': 'Complaint not found'}), 404
        
        citizen_id, is_valid = result
        
        # Update complaint
        cursor.execute('''
            UPDATE complaints 
            SET evidence = ?, status = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (evidence, status, complaint_id))
        
        # Update CRS based on resolution
        if status == 'Resolved' and is_valid:
            update_citizen_crs(citizen_id, True, False)
        elif status == 'Rejected' and is_valid:
            update_citizen_crs(citizen_id, False, False)
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'message': 'Complaint updated successfully',
            'complaint_id': complaint_id,
            'status': status
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

@app.route('/dashboard', methods=['GET'])
def get_dashboard():
    """Get dashboard data with complaint statistics"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Get all complaints with pagination
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        offset = (page - 1) * limit
        
        cursor.execute('''
            SELECT c.id, c.text, c.category, c.urgency, c.citizen_id, 
                   c.crs_score, c.hash, c.timestamp, c.status, c.evidence,
                   c.is_duplicate, c.is_valid
            FROM complaints c
            ORDER BY c.timestamp DESC
            LIMIT ? OFFSET ?
        ''', (limit, offset))
        
        complaints = cursor.fetchall()
        
        # Get statistics
        cursor.execute('SELECT COUNT(*) FROM complaints')
        total_complaints = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM complaints WHERE status = "Pending"')
        pending_complaints = cursor.fetchone()[0]
        
        cursor.execute('SELECT COUNT(*) FROM complaints WHERE is_duplicate = 1')
        duplicate_complaints = cursor.fetchone()[0]
        
        cursor.execute('SELECT AVG(crs_score) FROM citizens')
        avg_crs = cursor.fetchone()[0] or 0
        
        conn.close()
        
        # Format complaints data
        complaints_data = []
        for complaint in complaints:
            complaints_data.append({
                'id': complaint[0],
                'text': complaint[1],
                'category': complaint[2],
                'urgency': complaint[3],
                'citizen_id': complaint[4],
                'crs_score': complaint[5],
                'hash': complaint[6],
                'timestamp': complaint[7],
                'status': complaint[8],
                'evidence': complaint[9],
                'is_duplicate': bool(complaint[10]),
                'is_valid': bool(complaint[11])
            })
        
        return jsonify({
            'complaints': complaints_data,
            'statistics': {
                'total_complaints': total_complaints,
                'pending_complaints': pending_complaints,
                'duplicate_complaints': duplicate_complaints,
                'average_crs': round(avg_crs, 2)
            },
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total_complaints
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

@app.route('/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'models_loaded': zero_shot_classifier is not None and sentence_model is not None
    }), 200

def allowed_audio_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route(f'/api/{API_VERSION}/voice/upload', methods=['POST'])
def upload_voice_complaint():
    """
    Upload and process voice complaint
    
    Form data:
        - audio: Audio file (WAV, MP3, OGG, M4A, FLAC, WEBM)
        - citizen_id: Citizen ID
        - language: (optional) Language code (hi, ta, gu, etc.)
    
    Returns:
        JSON with complaint data including transcription
    """
    try:
        global voice_service
        
        # Check if audio file is present
        if 'audio' not in request.files:
            return jsonify({'status': 'error', 'message': 'No audio file provided'}), 400
        
        audio_file = request.files['audio']
        
        # Check if file was selected
        if audio_file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        # Validate file type
        if not allowed_audio_file(audio_file.filename):
            return jsonify({
                'status': 'error', 
                'message': f'Invalid file type. Allowed: {ALLOWED_EXTENSIONS}'
            }), 400
        
        # Get form data
        citizen_id = request.form.get('citizen_id')
        language = request.form.get('language', None)
        
        if not citizen_id:
            return jsonify({'status': 'error', 'message': 'citizen_id is required'}), 400
        
        # Save uploaded file temporarily
        filename = secure_filename(audio_file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{citizen_id}_{timestamp}_{filename}"
        file_path = os.path.join(UPLOAD_FOLDER, unique_filename)
        
        # Save file
        audio_file.save(file_path)
        logger.info(f"Audio file saved: {file_path}")
        
        try:
            # Initialize voice service if not already loaded
            if voice_service is None:
                logger.info("Loading voice service...")
                voice_service = get_voice_service(model_size='base')
            
            # Process voice complaint
            result = voice_service.process_voice_complaint(
                file_path=file_path,
                citizen_id=citizen_id,
                language=language
            )
            
            if not result['success']:
                return jsonify({
                    'status': 'error',
                    'message': result.get('error', 'Voice processing failed')
                }), 500
            
            # Get complaint text from transcription
            complaint_text = result['text']
            detected_language = result['language']
            
            # Classify complaint using multilingual classifier
            if multilingual_classifier:
                analysis = multilingual_classifier.analyze_complaint(
                    complaint_text, 
                    detected_language
                )
                category = analysis['category']
                urgency = analysis['urgency']
                keywords = analysis['keywords']
            else:
                # Fallback to default classifier
                category = classify_complaint(complaint_text)
                urgency = detect_urgency(complaint_text)
                keywords = []
            
            # Validate complaint context
            is_valid_context, validation_message = validate_complaint_context(complaint_text)
            
            # Check for duplicates
            is_duplicate, duplicate_id = detect_duplicates(complaint_text, citizen_id)
            
            # Generate blockchain hash
            complaint_hash = generate_blockchain_hash(complaint_text, result['timestamp'])
            
            # Update CRS score
            crs_score = update_citizen_crs(citizen_id, is_valid_context, is_duplicate)
            
            # Determine status
            if not is_valid_context:
                status = 'Invalid'
            elif is_duplicate:
                status = 'Duplicate'
            else:
                status = 'Pending'
            
            # Save complaint to database
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO complaints 
                (text, category, urgency, citizen_id, crs_score, hash, timestamp, 
                 status, is_duplicate, is_valid)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (complaint_text, category, urgency, citizen_id, crs_score, 
                  complaint_hash, result['timestamp'], status, is_duplicate, is_valid_context))
            
            complaint_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            # Prepare response in requested format
            response_data = {
                'complaint_id': f"GSAI-{datetime.now().year}-{str(complaint_id).zfill(4)}",
                'text': complaint_text,
                'category': category,
                'urgency': urgency,
                'language': detected_language,
                'timestamp': result['timestamp'],
                'hash': complaint_hash,
                'is_valid': is_valid_context,
                'is_duplicate': is_duplicate,
                'crs_score': crs_score,
                'audio_duration': result.get('audio_duration', 0),
                'keywords': keywords
            }
            
            if is_duplicate and duplicate_id:
                response_data['duplicate_of'] = duplicate_id
            
            logger.info(f"Voice complaint processed: {response_data['complaint_id']}")
            
            return jsonify({
                'status': 'success',
                'data': response_data
            }), 201
            
        finally:
            # Clean up uploaded file
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    logger.info(f"Cleaned up audio file: {file_path}")
                except Exception as e:
                    logger.warning(f"Failed to remove temp file: {e}")
        
    except Exception as e:
        logger.error(f"Voice complaint upload error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/voice/languages', methods=['GET'])
def get_supported_languages():
    """
    Get list of supported languages for voice input
    
    Returns:
        JSON with supported languages
    """
    try:
        global voice_service
        
        # Initialize voice service if needed
        if voice_service is None:
            voice_service = get_voice_service(model_size='base')
        
        languages = voice_service.get_supported_languages()
        
        return jsonify({
            'status': 'success',
            'data': {
                'languages': languages,
                'count': len(languages)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Error fetching languages: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route(f'/api/{API_VERSION}/voice/test', methods=['GET'])
def test_voice_service():
    """
    Test if voice service is available
    
    Returns:
        JSON with service status
    """
    try:
        global voice_service
        
        if voice_service is None:
            return jsonify({
                'status': 'success',
                'data': {
                    'available': False,
                    'message': 'Voice service not initialized. It will load on first use.'
                }
            }), 200
        
        return jsonify({
            'status': 'success',
            'data': {
                'available': True,
                'model_size': voice_service.model_size,
                'supported_formats': list(voice_service.SUPPORTED_FORMATS)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Voice service test error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route(f'/api/{API_VERSION}/sms/complaint', methods=['POST'])
def submit_sms_complaint():
    """
    Submit complaint via SMS
    
    Expected JSON:
    {
        "sms_content": "156Complaint#Water supply issue in sector 5",
        "sender_number": "+919876543210"
    }
    
    Returns:
        JSON with complaint processing results
    """
    try:
        from services.ivr_sms_service import get_ivr_sms_service
        
        data = request.get_json()
        
        if not data or 'sms_content' not in data or 'sender_number' not in data:
            return jsonify({
                'status': 'error',
                'message': 'Missing required fields: sms_content, sender_number'
            }), 400
        
        sms_content = data['sms_content']
        sender_number = data['sender_number']
        
        # Get IVR/SMS service
        ivr_sms_service = get_ivr_sms_service()
        
        # Process SMS complaint
        result = ivr_sms_service.process_sms_complaint(sms_content, sender_number)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 201
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 400
            
    except Exception as e:
        logger.error(f"SMS complaint processing error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/ussd/complaint', methods=['POST'])
def submit_ussd_complaint():
    """
    Submit complaint via USSD
    
    Expected JSON:
    {
        "ussd_input": "Water#No water supply in sector 5#Sector 5",
        "session_id": "USSD123456",
        "sender_number": "+919876543210"
    }
    
    Returns:
        JSON with complaint processing results
    """
    try:
        from services.ivr_sms_service import get_ivr_sms_service
        
        data = request.get_json()
        
        if not data or 'ussd_input' not in data or 'session_id' not in data or 'sender_number' not in data:
            return jsonify({
                'status': 'error',
                'message': 'Missing required fields: ussd_input, session_id, sender_number'
            }), 400
        
        ussd_input = data['ussd_input']
        session_id = data['session_id']
        sender_number = data['sender_number']
        
        # Get IVR/SMS service
        ivr_sms_service = get_ivr_sms_service()
        
        # Process USSD complaint
        result = ivr_sms_service.process_ussd_complaint(ussd_input, session_id, sender_number)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 201
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 400
            
    except Exception as e:
        logger.error(f"USSD complaint processing error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/csc/register', methods=['POST'])
def register_csc():
    """
    Register a new CSC
    
    Expected JSON:
    {
        "name": "CSC Sector 5",
        "location": "Main Market, Sector 5",
        "district": "South Delhi",
        "state": "Delhi",
        "pincode": "110010",
        "contact_person": "Rajesh Kumar",
        "phone": "+919876543210",
        "email": "csc5@example.com"
    }
    
    Returns:
        JSON with registration results
    """
    try:
        from services.csc_agent_service import get_csc_agent_service
        
        data = request.get_json()
        
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data provided'
            }), 400
        
        # Get CSC/Agent service
        csc_service = get_csc_agent_service()
        
        # Register CSC
        result = csc_service.register_csc(data)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 201
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 400
            
    except Exception as e:
        logger.error(f"CSC registration error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/agent/register', methods=['POST'])
def register_agent():
    """
    Register a new agent
    
    Expected JSON:
    {
        "csc_id": 1,
        "name": "Priya Sharma",
        "phone": "+919876543211",
        "email": "priya@csc5.com",
        "employee_id": "CSC5-AGENT001",
        "password": "securepassword",
        "role": "agent"
    }
    
    Returns:
        JSON with registration results
    """
    try:
        from services.csc_agent_service import get_csc_agent_service
        
        data = request.get_json()
        
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data provided'
            }), 400
        
        # Get CSC/Agent service
        csc_service = get_csc_agent_service()
        
        # Register agent
        result = csc_service.register_agent(data)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 201
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 400
            
    except Exception as e:
        logger.error(f"Agent registration error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/agent/login', methods=['POST'])
def login_agent():
    """
    Authenticate an agent
    
    Expected JSON:
    {
        "employee_id": "CSC5-AGENT001",
        "password": "securepassword"
    }
    
    Returns:
        JSON with authentication results and session token
    """
    try:
        from services.csc_agent_service import get_csc_agent_service
        
        data = request.get_json()
        
        if not data or 'employee_id' not in data or 'password' not in data:
            return jsonify({
                'status': 'error',
                'message': 'Missing required fields: employee_id, password'
            }), 400
        
        employee_id = data['employee_id']
        password = data['password']
        
        # Get CSC/Agent service
        csc_service = get_csc_agent_service()
        
        # Authenticate agent
        result = csc_service.authenticate_agent(employee_id, password)
        
        if result['success']:
            # Start kiosk session
            session_result = csc_service.start_kiosk_session(
                result['agent']['csc_id'],
                result['agent']['id'],
                request.remote_addr,
                request.headers.get('User-Agent')
            )
            
            if session_result['success']:
                response_data = {
                    'agent': result['agent'],
                    'session_token': session_result['session_token']
                }
                return jsonify({
                    'status': 'success',
                    'data': response_data
                }), 200
            else:
                return jsonify({
                    'status': 'error',
                    'message': session_result['error']
                }), 500
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 401
            
    except Exception as e:
        logger.error(f"Agent login error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/csc/<int:csc_id>', methods=['GET'])
def get_csc_info(csc_id):
    """
    Get CSC information
    
    Args:
        csc_id: CSC ID
        
    Returns:
        JSON with CSC information
    """
    try:
        from services.csc_agent_service import get_csc_agent_service
        
        # Get CSC/Agent service
        csc_service = get_csc_agent_service()
        
        # Get CSC info
        result = csc_service.get_csc_info(csc_id)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result['csc']
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 404
            
    except Exception as e:
        logger.error(f"CSC info retrieval error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/csc/<int:csc_id>/agents', methods=['GET'])
def get_csc_agents(csc_id):
    """
    Get agents for a CSC
    
    Args:
        csc_id: CSC ID
        
    Returns:
        JSON with agents list
    """
    try:
        from services.csc_agent_service import get_csc_agent_service
        
        # Get CSC/Agent service
        csc_service = get_csc_agent_service()
        
        # Get CSC agents
        result = csc_service.get_csc_agents(csc_id)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result['agents']
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 404
            
    except Exception as e:
        logger.error(f"CSC agents retrieval error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/fraud/check', methods=['POST'])
def check_fraud_risk():
    """
    Check fraud risk for a complaint
    
    Expected JSON:
    {
        "text": "Complaint description",
        "citizen_id": "CITIZEN001",
        "category": "Water",
        "urgency": "Medium"
    }
    
    Returns:
        JSON with fraud risk assessment
    """
    try:
        from services.fraud_detection_service import get_fraud_detection_service
        
        data = request.get_json()
        
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data provided'
            }), 400
        
        # Get fraud detection service
        fraud_service = get_fraud_detection_service()
        
        # Check fraud risk
        result = fraud_service.detect_fraud_risk(data)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Fraud risk check error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/complaints/duplicates', methods=['POST'])
def check_duplicates():
    """
    Check for duplicate complaints
    
    Expected JSON:
    {
        "text": "Complaint description",
        "citizen_id": "CITIZEN001"
    }
    
    Returns:
        JSON with duplicate detection results
    """
    try:
        from services.fraud_detection_service import get_fraud_detection_service
        
        data = request.get_json()
        
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data provided'
            }), 400
        
        # Get fraud detection service
        fraud_service = get_fraud_detection_service()
        
        # Check for duplicates
        result = fraud_service.detect_duplicates_advanced(data)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Duplicate check error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/audit/log', methods=['POST'])
def log_audit_trail_event():
    """
    Log an audit event
    
    Expected JSON:
    {
        "event_type": "complaint_submitted",
        "entity_type": "complaint",
        "entity_id": "COMPLAINT001",
        "action": "submit",
        "actor_id": "CITIZEN001",
        "actor_role": "citizen",
        "data": {"category": "Water", "urgency": "High"}
    }
    
    Returns:
        JSON with log result
    """
    try:
        from services.audit_service import get_audit_service
        
        data = request.get_json()
        
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No data provided'
            }), 400
        
        # Get audit service
        audit_service = get_audit_service()
        
        # Log event
        result = audit_service.log_event(data)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 201
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Audit event logging error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/audit/verify', methods=['GET'])
def verify_audit_trail():
    """
    Verify the integrity of the audit trail
    
    Returns:
        JSON with verification results
    """
    try:
        from services.audit_service import get_audit_service
        
        # Get audit service
        audit_service = get_audit_service()
        
        # Verify audit trail
        result = audit_service.verify_audit_trail()
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Audit trail verification error: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/audit/trail', methods=['GET'])
def get_audit_trail_events():
    """
    Get audit trail
    
    Query Parameters:
        entity_type: Type of entity (optional)
        entity_id: ID of entity (optional)
        limit: Maximum number of events (default: 100)
    
    Returns:
        JSON with audit trail
    """
    try:
        from services.audit_service import get_audit_service
        
        # Get parameters
        entity_type = request.args.get('entity_type')
        entity_id = request.args.get('entity_id')
        limit = int(request.args.get('limit', 100))
        
        # Get audit service
        audit_service = get_audit_service()
        
        # Get audit trail
        result = audit_service.get_audit_trail(entity_type, entity_id, limit)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Error retrieving audit trail: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/analytics/trends', methods=['GET'])
def get_complaint_trends():
    """
    Get complaint trends over time
    
    Query Parameters:
        days: Number of days to analyze (default: 30)
    
    Returns:
        JSON with trend data
    """
    try:
        from services.analytics_service import get_analytics_service
        
        # Get parameters
        days = int(request.args.get('days', 30))
        
        # Get analytics service
        analytics_service = get_analytics_service()
        
        # Get trends
        result = analytics_service.get_complaint_trends(days)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result['data']
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Error getting complaint trends: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/analytics/categories', methods=['GET'])
def get_category_analysis():
    """
    Get detailed category analysis
    
    Returns:
        JSON with category analysis
    """
    try:
        from services.analytics_service import get_analytics_service
        
        # Get analytics service
        analytics_service = get_analytics_service()
        
        # Get category analysis
        result = analytics_service.get_category_analysis()
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result['data']
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Error getting category analysis: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/analytics/heatmap', methods=['GET'])
def get_heatmap_data():
    """
    Get heatmap data for complaint hotspots
    
    Returns:
        JSON with heatmap data
    """
    try:
        from services.analytics_service import get_analytics_service
        
        # Get analytics service
        analytics_service = get_analytics_service()
        
        # Get heatmap data
        result = analytics_service.get_heatmap_data()
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result['data'],
                'categories': result['categories']
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Error getting heatmap data: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/analytics/sentiment', methods=['GET'])
def get_sentiment_analysis():
    """
    Get sentiment analysis of complaints
    
    Returns:
        JSON with sentiment analysis
    """
    try:
        from services.analytics_service import get_analytics_service
        
        # Get analytics service
        analytics_service = get_analytics_service()
        
        # Get sentiment analysis
        result = analytics_service.get_sentiment_analysis()
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result['data']
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Error getting sentiment analysis: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/analytics/resources', methods=['GET'])
def get_resource_allocation_insights():
    """
    Get insights for resource allocation
    
    Returns:
        JSON with resource allocation insights
    """
    try:
        from services.analytics_service import get_analytics_service
        
        # Get analytics service
        analytics_service = get_analytics_service()
        
        # Get resource allocation insights
        result = analytics_service.get_resource_allocation_insights()
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result['data']
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Error getting resource allocation insights: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route(f'/api/{API_VERSION}/analytics/predict', methods=['GET'])
def predict_complaint_volume():
    """
    Predict complaint volume
    
    Query Parameters:
        days: Number of days to forecast (default: 7)
    
    Returns:
        JSON with predictions
    """
    try:
        from services.analytics_service import get_analytics_service
        
        # Get parameters
        days = int(request.args.get('days', 7))
        
        # Get analytics service
        analytics_service = get_analytics_service()
        
        # Get predictions
        result = analytics_service.predict_complaint_volume(days)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result['data']
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Error predicting complaint volume: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

        audit_service = get_audit_service()
        
        # Get audit trail
        result = audit_service.get_audit_trail(entity_type, entity_id, limit)
        
        if result['success']:
            return jsonify({
                'status': 'success',
                'data': result
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': result['error']
            }), 500
            
    except Exception as e:
        logger.error(f"Error retrieving audit trail: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Internal server error: {str(e)}'
        }), 500

@app.route('/citizen/<citizen_id>', methods=['GET'])
def get_citizen_info(citizen_id):
    """Get citizen information and CRS score"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM citizens WHERE id = ?', (citizen_id,))
        citizen = cursor.fetchone()
        
        if not citizen:
            conn.close()
            return jsonify({'error': 'Citizen not found'}), 404
        
        # Get complaint history
        cursor.execute('''
            SELECT id, category, status, timestamp, is_valid
            FROM complaints 
            WHERE citizen_id = ?
            ORDER BY timestamp DESC
            LIMIT 10
        ''', (citizen_id,))
        
        complaints = cursor.fetchall()
        conn.close()
        
        return jsonify({
            'citizen_id': citizen[0],
            'crs_score': citizen[1],
            'created_at': citizen[2],
            'updated_at': citizen[3],
            'recent_complaints': [
                {
                    'id': c[0],
                    'category': c[1],
                    'status': c[2],
                    'timestamp': c[3],
                    'is_valid': bool(c[4])
                }
                for c in complaints
            ]
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

# ============================================
# EXPORT ENDPOINTS (PDF, Excel, Email)
# ============================================

@app.route('/api/export/pdf', methods=['POST'])
def export_pdf():
    """
    Export dashboard data as PDF
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        
        data = request.get_json()
        role = data.get('role', 'citizen')
        time_range = data.get('timeRange', '30d')
        
        # Create PDF in memory
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#667eea'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        elements.append(Paragraph(f"GramSetu AI - {role.title()} Dashboard Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Metadata
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Paragraph(f"Time Range: {time_range}", styles['Normal']))
        elements.append(Paragraph(f"Report Type: {role.title()} Analytics", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Get dashboard data
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # Summary statistics
        cursor.execute("SELECT COUNT(*) FROM complaints")
        total = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM complaints WHERE status = 'Resolved'")
        resolved = cursor.fetchone()[0]
        
        cursor.execute("SELECT COUNT(*) FROM complaints WHERE status = 'Pending'")
        pending = cursor.fetchone()[0]
        
        # KPI Table
        kpi_data = [
            ['Metric', 'Value', 'Status'],
            ['Total Complaints', str(total), '✓'],
            ['Resolved', str(resolved), f"{(resolved/total*100):.1f}%" if total > 0 else '0%'],
            ['Pending', str(pending), f"{(pending/total*100):.1f}%" if total > 0 else '0%'],
            ['Resolution Rate', f"{(resolved/total*100):.1f}%" if total > 0 else '0%', '✓']
        ]
        
        kpi_table = Table(kpi_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(kpi_table)
        elements.append(Spacer(1, 0.5*inch))
        
        # Recent complaints
        elements.append(Paragraph("Recent Complaints", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        cursor.execute("""
            SELECT id, category, status, urgency, timestamp 
            FROM complaints 
            ORDER BY timestamp DESC 
            LIMIT 10
        """)
        
        complaints_data = [['ID', 'Category', 'Status', 'Urgency', 'Date']]
        for row in cursor.fetchall():
            complaints_data.append([
                f"C{row[0]}",
                row[1] or 'N/A',
                row[2],
                row[3],
                row[4][:10] if row[4] else 'N/A'
            ])
        
        complaints_table = Table(complaints_data, colWidths=[0.8*inch, 1.5*inch, 1.2*inch, 1*inch, 1.5*inch])
        complaints_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey)
        ]))
        
        elements.append(complaints_table)
        conn.close()
        
        # Footer
        elements.append(Spacer(1, 1*inch))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        elements.append(Paragraph("GramSetu AI - National Governance Intelligence Network", footer_style))
        elements.append(Paragraph("Powered by AI | Secured by Blockchain | Digital India Initiative", footer_style))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'GramSetu_{role}_Report_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
        
    except ImportError:
        logger.error("reportlab not installed")
        return jsonify({"status": "error", "message": "PDF export requires reportlab package"}), 500
    except Exception as e:
        logger.error(f"Error exporting PDF: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """
    Export dashboard data as Excel
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        data = request.get_json()
        role = data.get('role', 'citizen')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"{role.title()} Dashboard"
        
        # Header styling
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        # Headers
        headers = ['Complaint ID', 'Category', 'Status', 'Urgency', 'Created Date', 'Is Valid', 'Is Duplicate']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Get data
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, category, status, urgency, timestamp, is_valid, is_duplicate 
            FROM complaints 
            ORDER BY timestamp DESC
        """)
        
        for row in cursor.fetchall():
            ws.append([
                f"C{row[0]}",
                row[1] or 'N/A',
                row[2],
                row[3],
                row[4][:19] if row[4] else 'N/A',
                'Yes' if row[5] else 'No',
                'Yes' if row[6] else 'No'
            ])
        
        conn.close()
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'GramSetu_{role}_Data_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except ImportError:
        logger.error("openpyxl not installed")
        return jsonify({"status": "error", "message": "Excel export requires openpyxl package"}), 500
    except Exception as e:
        logger.error(f"Error exporting Excel: {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    # Initialize database
    init_database()
    
    # Initialize AI models
    initialize_ai_models()
    
    # Start Flask server
    port = int(os.getenv('PORT', 5001))
    logger.info(f"Starting GramSetu AI backend server on http://localhost:{port}")
    app.run(host='0.0.0.0', port=port, debug=True)


